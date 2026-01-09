import pandas as pd
import numpy as np
from datetime import datetime

# ============================================
# ГЕНЕРАЦИЯ ТЕСТОВЫХ ДАННЫХ (если файл не найден)
# ============================================

def generate_test_data():
    """Генерация тестовых данных журнала успеваемости"""
    np.random.seed(42)
    students = [f'Ученик_{i}' for i in range(1, 26)]
    subjects = ['Математика', 'Теорема Пифагора', 'Русский', 'Физика', 'Информатика']
    
    data = {'Ученик': students}
    for subject in subjects:
        data[subject] = np.random.randint(2, 6, len(students))
    
    df = pd.DataFrame(data)
    print("✅ Тестовые данные сгенерированы")
    return df

# ============================================
# ЗАГРУЗКА И ПОДГОТОВКА ДАННЫХ
# ============================================

def load_journal(filename='journal.csv'):
    """Загрузка журнала из файла CSV"""
    try:
        df = pd.read_csv(filename, encoding='utf-8')
        print(f"✅ Данные успешно загружены из {filename}")
        print(f"   Количество учеников: {len(df)}")
        return df
    except FileNotFoundError:
        print(f"❌ Файл {filename} не найден! Генерация тестовых данных...")
        return generate_test_data()
    except Exception as e:
        print(f"❌ Ошибка загрузки: {e}")
        return None

# ============================================
# АНАЛИЗ ДАННЫХ
# ============================================

def calculate_statistics(df):
    """Расчёт статистики по журналу"""
    if df is None:
        return None, None
    
    # Определяем столбцы с предметами
    subject_columns = [col for col in df.columns if col != 'Ученик']
    
    # Расчёт среднего балла каждого ученика
    df['Средний_балл'] = df[subject_columns].mean(axis=1).round(2)
    
    # Определение статуса
    def get_status(avg):
        if avg >= 4.5:
            return 'Отличник'
        elif avg >= 3.5:
            return 'Хорошист'
        elif avg >= 2.5:
            return 'Троечник'
        else:
            return 'Требует внимания'
    
    df['Статус'] = df['Средний_балл'].apply(get_status)
    
    return df, subject_columns

def get_class_statistics(df, subject_columns):
    """Получение общей статистики класса"""
    if df is None:
        return None
    
    stats = {
        'total_students': len(df),
        'class_average': df['Средний_балл'].mean().round(2),
        'class_median': df['Средний_балл'].median().round(2),
        'class_std': df['Средний_балл'].std().round(2),
        'class_min': df['Средний_балл'].min(),
        'class_max': df['Средний_балл'].max(),
        'excellent': len(df[df['Статус'] == 'Отличник']),
        'good': len(df[df['Статус'] == 'Хорошист']),
        'satisfactory': len(df[df['Статус'] == 'Троечник']),
        'attention_needed': len(df[df['Статус'] == 'Требует внимания'])
    }
    
    subject_stats = {}
    for subject in subject_columns:
        subject_stats[subject] = {
            'mean': df[subject].mean().round(2),
            'min': df[subject].min(),
            'max': df[subject].max()
        }
    
    return stats, subject_stats

def get_top_students(df, n=5):
    """Получение топ-N лучших учеников"""
    return df.sort_values('Средний_балл', ascending=False).head(n)

def get_struggling_students(df):
    """Выявление отстающих учеников (средний балл < 3.5)"""
    return df[df['Средний_балл'] < 3.5].sort_values('Средний_балл')

# ============================================
# СОХРАНЕНИЕ РЕЗУЛЬТАТОВ
# ============================================

def save_to_excel(df, filename='journal_analysis.xlsx'):
    """Сохранение результатов в Excel с форматированием"""
    if df is None:
        return
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Журнал', index=False)
        
        # Форматирование
        worksheet = writer.sheets['Журнал']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем условное форматирование для статуса
        from openpyxl.styles import PatternFill
        excellent_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Зеленый
        good_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Желтый
        satisfactory_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Оранжевый
        attention_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Красный
        
        status_col = df.columns.get_loc('Статус') + 1  # +1 для Excel
        for row in range(2, len(df) + 2):  # Начиная со второй строки
            status = worksheet.cell(row=row, column=status_col).value
            if status == 'Отличник':
                worksheet.cell(row=row, column=status_col).fill = excellent_fill
            elif status == 'Хорошист':
                worksheet.cell(row=row, column=status_col).fill = good_fill
            elif status == 'Троечник':
                worksheet.cell(row=row, column=status_col).fill = satisfactory_fill
            elif status == 'Требует внимания':
                worksheet.cell(row=row, column=status_col).fill = attention_fill
    
    print(f"✅ Результаты сохранены в {filename}")

def create_text_report(df, stats, subject_stats, filename='report.txt'):
    """Создание текстового отчёта"""
    if df is None:
        return
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("ОТЧЁТ ПО ЖУРНАЛУ УСПЕВАЕМОСТИ\n")
        f.write("=" * 80 + "\n")
        f.write(f"Дата отчёта: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n")
        
        f.write("1. ОБЩАЯ СТАТИСТИКА КЛАССА\n")
        f.write("-" * 80 + "\n")
        f.write(f"Всего учеников: {stats['total_students']}\n")
        f.write(f"Средний балл класса: {stats['class_average']:.2f}\n")
        f.write(f"Медиана: {stats['class_median']:.2f}\n")
        f.write(f"Стандартное отклонение: {stats['class_std']:.2f}\n")
        f.write(f"Минимальный балл: {stats['class_min']:.2f}\n")
        f.write(f"Максимальный балл: {stats['class_max']:.2f}\n\n")
        
        f.write(f"Распределение статусов:\n")
        f.write(f"  Отличников: {stats['excellent']}\n")
        f.write(f"  Хорошистов: {stats['good']}\n")
        f.write(f"  Троечников: {stats['satisfactory']}\n")
        f.write(f"  Требуют внимания: {stats['attention_needed']}\n\n")
        
        f.write("2. ТОП-5 ЛУЧШИХ УЧЕНИКОВ\n")
        f.write("-" * 80 + "\n")
        top = get_top_students(df, 5)
        for i, (_, row) in enumerate(top.iterrows(), 1):
            f.write(f"{i}. {row['Ученик']}: {row['Средний_балл']:.2f} ({row['Статус']})\n")
        
        f.write("\n3. УЧЕНИКИ, ТРЕБУЮЩИЕ ВНИМАНИЯ\n")
        f.write("-" * 80 + "\n")
        struggling = get_struggling_students(df)
        if len(struggling) > 0:
            for _, row in struggling.iterrows():
                f.write(f"- {row['Ученик']}: {row['Средний_балл']:.2f} ({row['Статус']})\n")
        else:
            f.write("Нет учеников с баллом ниже 3.5\n")
        
        f.write("\n4. СТАТИСТИКА ПО ПРЕДМЕТАМ\n")
        f.write("-" * 80 + "\n")
        for subject, s_stats in subject_stats.items():
            f.write(f"{subject}:\n")
            f.write(f"  Среднее: {s_stats['mean']:.2f}\n")
            f.write(f"  Минимум: {s_stats['min']}\n")
            f.write(f"  Максимум: {s_stats['max']}\n\n")
        
        f.write("=" * 80 + "\n")
    
    print(f"✅ Отчёт сохранён в {filename}")

# ============================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================

def main():
    """Главная функция анализа журнала"""
    
    print("\n" + "=" * 80)
    print("АНАЛИЗ ЖУРНАЛА УСПЕВАЕМОСТИ")
    print("=" * 80 + "\n")
    
    # Загрузка данных
    df = load_journal()
    if df is None:
        return
    
    # Расчёт статистики
    df, subject_columns = calculate_statistics(df)
    
    # Общая статистика
    stats, subject_stats = get_class_statistics(df, subject_columns)
    
    # Вывод результатов
    print("\n" + "=" * 80)
    print("РЕЗУЛЬТАТЫ АНАЛИЗА")
    print("=" * 80)
    
    print(f"\n📈 Средний балл класса: {stats['class_average']:.2f}")
    print(f"📊 Медиана: {stats['class_median']:.2f}")
    print(f"📉 Стандартное отклонение: {stats['class_std']:.2f}")
    
    print(f"\n👥 Распределение учеников:")
    print(f"   Отличников: {stats['excellent']}")
    print(f"   Хорошистов: {stats['good']}")
    print(f"   Троечников: {stats['satisfactory']}")
    print(f"   Требуют внимания: {stats['attention_needed']}")
    
    print("\n🏆 ТОП-5 ЛУЧШИХ УЧЕНИКОВ:")
    top = get_top_students(df, 5)
    for i, (_, row) in enumerate(top.iterrows(), 1):
        print(f"   {i}. {row['Ученик']}: {row['Средний_балл']:.2f}")
    
    struggling = get_struggling_students(df)
    if len(struggling) > 0:
        print(f"\n⚠️  УЧЕНИКИ, ТРЕБУЮЩИЕ ВНИМАНИЯ ({len(struggling)} чел.):")
        for _, row in struggling.iterrows():
            print(f"   • {row['Ученик']}: {row['Средний_балл']:.2f}")
    
    print("\n📚 СТАТИСТИКА ПО ПРЕДМЕТАМ:")
    for subject, s_stats in subject_stats.items():
        print(f"   {subject}: среднее = {s_stats['mean']:.2f}, "
              f"мин = {s_stats['min']:.0f}, макс = {s_stats['max']:.0f}")
    
    # Самый сложный предмет
    easiest_subject = min(subject_stats.items(), key=lambda x: x[1]['mean'])
    hardest_subject = max(subject_stats.items(), key=lambda x: x[1]['mean'])
    print(f"\n   🎯 Лучший предмет: {hardest_subject[0]} ({hardest_subject[1]['mean']:.2f})")
    print(f"   ⚠️  Сложный предмет: {easiest_subject[0]} ({easiest_subject[1]['mean']:.2f})")
    
    # Сохранение результатов
    print("\n💾 Сохранение результатов...")
    save_to_excel(df, 'journal_analysis.xlsx')
    create_text_report(df, stats, subject_stats, 'report.txt')
    
    print("\n✅ Анализ завершён!")
    print("=" * 80 + "\n")

if __name__ == "__main__":
    main()